"""Apply reviewer decisions from a returned review workbook.

  APPROVE = YES  -> status 'live'
  APPROVE = NO   -> status 'retired', store REVIEW_NOTES
  APPROVE = EDIT -> apply CORRECTED_ANSWER to answer_key.correct, status 'live', store notes
  blank          -> left as 'draft', counted as unreviewed
"""
from __future__ import annotations

import json

from openpyxl import load_workbook

from . import db


def _norm(v) -> str:
    return ("" if v is None else str(v)).strip()


def import_review(path: str, dry_run: bool = False) -> dict:
    wb = load_workbook(path)
    ws = wb["Review"] if "Review" in wb.sheetnames else wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: i for i, name in enumerate(header_cells)}
    for required in ("mission_id", "APPROVE"):
        if required not in headers:
            raise ValueError(f"review file is missing required column '{required}'")

    counts = {"approved": 0, "rejected": 0, "edited": 0, "unreviewed": 0, "skipped": 0}

    conn = db.get_connection()
    try:
        cur = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):
            mid_raw = row[headers["mission_id"]]
            if mid_raw in (None, ""):
                continue
            mission_id = int(mid_raw)
            decision = _norm(row[headers["APPROVE"]]).upper()
            notes = _norm(row[headers["REVIEW_NOTES"]]) if "REVIEW_NOTES" in headers else ""
            corrected = _norm(row[headers["CORRECTED_ANSWER"]]) if "CORRECTED_ANSWER" in headers else ""

            if decision == "YES":
                counts["approved"] += 1
                if not dry_run:
                    cur.execute(
                        "UPDATE missions SET status='live', review_notes=%s WHERE id=%s AND status='draft'",
                        (notes or None, mission_id),
                    )
            elif decision == "NO":
                counts["rejected"] += 1
                if not dry_run:
                    cur.execute(
                        "UPDATE missions SET status='retired', review_notes=%s WHERE id=%s AND status='draft'",
                        (notes or None, mission_id),
                    )
            elif decision == "EDIT":
                counts["edited"] += 1
                if not dry_run:
                    cur.execute("SELECT answer_key FROM missions WHERE id=%s AND status='draft'", (mission_id,))
                    r = cur.fetchone()
                    if r is None:
                        counts["edited"] -= 1
                        counts["skipped"] += 1
                        print(f"  mission {mission_id}: not found in 'draft' status; skipped.")
                        continue
                    ak = r[0]
                    if isinstance(ak, (bytes, bytearray)):
                        ak = ak.decode("utf-8")
                    ak = json.loads(ak) if isinstance(ak, str) else (ak or {})
                    if corrected:
                        ak["correct"] = corrected.lower()
                    cur.execute(
                        "UPDATE missions SET status='live', answer_key=%s, review_notes=%s WHERE id=%s",
                        (json.dumps(ak), notes or None, mission_id),
                    )
            else:
                counts["unreviewed"] += 1

        if not dry_run:
            conn.commit()
        cur.close()
    finally:
        conn.close()

    prefix = "  [dry-run] would apply" if dry_run else "  Review applied:"
    print(
        f"{prefix} {counts['approved']} approved, {counts['rejected']} rejected, "
        f"{counts['edited']} edited, {counts['unreviewed']} unreviewed"
        + (f", {counts['skipped']} skipped" if counts["skipped"] else "")
        + "."
    )
    return counts
