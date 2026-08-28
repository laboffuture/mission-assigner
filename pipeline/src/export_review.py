"""Export all status='draft' missions to an Excel review workbook.

This spreadsheet IS the review workspace for the pilot — there is no review UI.
The SME/QC fills the APPROVE column (dropdown YES/NO/EDIT), optionally
CORRECTED_ANSWER and REVIEW_NOTES, and returns the file to import_review.
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import db

HEADERS = [
    "mission_id", "source_file", "chunk_ref", "difficulty", "level_name", "tags",
    "title", "question", "option_a", "option_b", "option_c", "option_d",
    "correct", "explanation", "source_quote",
    "APPROVE", "CORRECTED_ANSWER", "REVIEW_NOTES",
]

# sensible display widths per column
WIDTHS = {
    "mission_id": 11, "source_file": 16, "chunk_ref": 22, "difficulty": 10,
    "level_name": 12, "tags": 16, "title": 24, "question": 50,
    "option_a": 30, "option_b": 30, "option_c": 30, "option_d": 30,
    "correct": 9, "explanation": 40, "source_quote": 50,
    "APPROVE": 12, "CORRECTED_ANSWER": 16, "REVIEW_NOTES": 30,
}


def _fetch_draft_rows():
    levels = {l["level"]: l["name"] for l in db.load_levels()["levels"]}
    conn = db.get_connection()
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """SELECT m.id, m.title, m.body, m.difficulty, m.answer_key,
                      c.source_file, c.chunk_ref
                 FROM missions m
                 LEFT JOIN content_chunks c ON c.id = m.source_chunk_id
                WHERE m.status = 'draft'
                ORDER BY c.source_file, c.chunk_ref, m.difficulty, m.id"""
        )
        missions = cur.fetchall()

        rows = []
        for m in missions:
            cur.execute(
                "SELECT option_key, option_text FROM mission_options WHERE mission_id = %s ORDER BY option_key",
                (m["id"],),
            )
            opts = {r["option_key"]: r["option_text"] for r in cur.fetchall()}
            cur.execute("SELECT tag FROM mission_tags WHERE mission_id = %s", (m["id"],))
            tags = ", ".join(r["tag"] for r in cur.fetchall())

            ak = m["answer_key"]
            if isinstance(ak, (bytes, bytearray)):
                ak = ak.decode("utf-8")
            ak = json.loads(ak) if isinstance(ak, str) else (ak or {})

            rows.append({
                "mission_id": m["id"],
                "source_file": m["source_file"] or "",
                "chunk_ref": m["chunk_ref"] or "",
                "difficulty": m["difficulty"],
                "level_name": levels.get(m["difficulty"], ""),
                "tags": tags,
                "title": m["title"],
                "question": m["body"],
                "option_a": opts.get("a", ""),
                "option_b": opts.get("b", ""),
                "option_c": opts.get("c", ""),
                "option_d": opts.get("d", ""),
                "correct": ak.get("correct", ""),
                "explanation": ak.get("explanation", ""),
                "source_quote": ak.get("source_quote", ""),
                "APPROVE": "",
                "CORRECTED_ANSWER": "",
                "REVIEW_NOTES": "",
            })
        cur.close()
        return rows
    finally:
        conn.close()


def export_review(out_path: str | None = None, dry_run: bool = False) -> str:
    rows = _fetch_draft_rows()
    if out_path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = str(db.PIPELINE_ROOT / f"review_{stamp}.xlsx")

    print(f"  {len(rows)} draft mission(s) to export.")
    if dry_run:
        print(f"  [dry-run] would write Excel review file to {out_path}")
        return out_path

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    review_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(HEADERS)
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = WIDTHS.get(name, 18)

    for r in rows:
        ws.append([r[h] for h in HEADERS])

    # wrap the long text columns
    wrap_cols = {"question", "explanation", "source_quote", "option_a", "option_b", "option_c", "option_d"}
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            header = HEADERS[cell.column - 1]
            if header in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # freeze the header row
    ws.freeze_panes = "A2"

    # data validation dropdown on APPROVE (YES/NO/EDIT)
    approve_col = get_column_letter(HEADERS.index("APPROVE") + 1)
    dv = DataValidation(type="list", formula1='"YES,NO,EDIT"', allow_blank=True)
    dv.prompt = "Choose YES, NO or EDIT"
    dv.promptTitle = "Review decision"
    ws.add_data_validation(dv)
    last_row = max(ws.max_row, 2)
    dv.add(f"{approve_col}2:{approve_col}{last_row}")
    # tint the reviewer-editable columns
    for name in ("APPROVE", "CORRECTED_ANSWER", "REVIEW_NOTES"):
        col = get_column_letter(HEADERS.index(name) + 1)
        for row in range(2, last_row + 1):
            ws[f"{col}{row}"].fill = review_fill

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"  Wrote review workbook: {out_path}")
    return out_path
